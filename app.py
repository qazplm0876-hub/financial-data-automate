import io, time, zipfile, xml.etree.ElementTree as ET, os
from datetime import datetime
import requests, pandas as pd, FinanceDataReader as fdr, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import Flask, request, send_file, jsonify

app = Flask(__name__)
DART_API_KEY = "fe92fa3b151b8990320a9fcfcfaac735101dbd07"

TARGET_ACCOUNTS = {
    "매출액":        ["매출액", "수익(매출액)", "영업수익", "매출", "순매출액"],
    "매출원가":      ["매출원가", "영업비용", "매출액의원가"],
    "판매비와관리비": ["판매비와관리비", "판매비와일반관리비", "판매관리비"],
    "영업이익":      ["영업이익", "영업이익(손실)", "영업손익"],
    "당기순이익":    ["당기순이익", "당기순이익(손실)", "분기순이익", "당기순손익"],
    "자산":          ["자산총계", "자산", "총자산"],
    "부채":          ["부채총계", "부채", "총부채"],
    "자본":          ["자본총계", "자본", "총자본"],
}
REPORT_FALLBACK = [
    ("2025", "11013", "2025년 1분기"),
    ("2024", "11011", "2024년 사업보고서"),
    ("2024", "11012", "2024년 반기"),
    ("2024", "11013", "2024년 1분기"),
]
_sl = None
_cm = None

def get_stock_list():
    global _sl
    if _sl is None:
        _sl = fdr.StockListing("KRX")
    return _sl

def build_corp_map():
    global _cm
    if _cm is not None:
        return _cm
    r = requests.get("https://opendart.fss.or.kr/api/corpCode.xml",
                     params={"crtfc_key": DART_API_KEY}, timeout=30)
    r.raise_for_status()
    with zipfile.ZipFile(io.BytesIO(r.content)) as z:
        with z.open("CORPCODE.xml") as f:
            tree = ET.parse(f)
    _cm = {}
    for item in tree.getroot().findall("list"):
        sc = item.findtext("stock_code", "").strip()
        if sc:
            _cm[sc] = (item.findtext("corp_code"), item.findtext("corp_name"))
    return _cm

def resolve_code(name, sl):
    exact = sl[sl["Name"].str.strip() == name.strip()]
    if len(exact) >= 1:
        return exact.iloc[0]["Code"], exact.iloc[0]["Name"], []
    partial = sl[sl["Name"].str.contains(name, na=False)]
    if partial.empty:
        return None, None, []
    if len(partial) == 1:
        return partial.iloc[0]["Code"], partial.iloc[0]["Name"], []
    return None, None, partial[["Code","Name","Market"]].to_dict("records")

def fetch_fs(corp_code, year, reprt_code):
    r = requests.get("https://opendart.fss.or.kr/api/fnlttSinglAcntAll.json",
                     params={"crtfc_key": DART_API_KEY, "corp_code": corp_code,
                             "bsns_year": year, "reprt_code": reprt_code, "fs_div": "OFS"},
                     timeout=15)
    r.raise_for_status()
    data = r.json()
    if data.get("status") != "000" or "list" not in data:
        return None
    return pd.DataFrame(data["list"])

def extract_account(df, aliases):
    for alias in aliases:
        matched = df[df["account_nm"].str.strip() == alias]
        if matched.empty:
            matched = df[df["account_nm"].str.contains(alias, na=False)]
        if not matched.empty:
            val = matched.iloc[0]["thstrm_amount"]
            if pd.notna(val) and str(val).strip() not in ("", "-", "N/A"):
                try:
                    return float(str(val).replace(",", ""))
                except ValueError:
                    pass
    return None

def query_one(stock_code, company_name, corp_map):
    if stock_code not in corp_map:
        return {"종목코드": stock_code, "기업명": company_name,
                "보고서기준": "조회실패", **{k: None for k in TARGET_ACCOUNTS}}
    corp_code, dart_name = corp_map[stock_code]
    display_name = dart_name or company_name
    df_fs, report_found = None, ""
    try:
        df_try = fetch_fs(corp_code, "2025", "11011")
        if df_try is not None and not df_try.empty:
            df_fs, report_found = df_try, "2025년 사업보고서"
    except Exception:
        pass
    if df_fs is None:
        for year, reprt_code, label in REPORT_FALLBACK:
            try:
                df_try = fetch_fs(corp_code, year, reprt_code)
                if df_try is not None and not df_try.empty:
                    df_fs, report_found = df_try, label
                    break
            except Exception:
                pass
            time.sleep(0.2)
    if df_fs is None:
        return {"종목코드": stock_code, "기업명": display_name,
                "보고서기준": "데이터없음", **{k: None for k in TARGET_ACCOUNTS}}
    row = {"종목코드": stock_code, "기업명": display_name, "보고서기준": report_found}
    for label, aliases in TARGET_ACCOUNTS.items():
        row[label] = extract_account(df_fs, aliases)
    return row

HTML = (
'<!doctype html><html lang="ko"><head>'
'<meta charset="utf-8">'
'<meta name="viewport" content="width=device-width,initial-scale=1">'
'<title>상장사 재무정보 조회</title>'
'<style>'
'*{box-sizing:border-box;margin:0;padding:0}'
'body{font-family:"Malgun Gothic",Arial,sans-serif;background:#f0f4f8;color:#222;font-size:14px}'
'header{background:#1F4E79;color:#fff;padding:20px 32px}'
'header h1{font-size:20px;font-weight:700;margin-bottom:4px}'
'header p{font-size:12px;color:#9ab8d4}'
'.wrap{max-width:1400px;margin:28px auto;padding:0 20px}'
'.card{background:#fff;border-radius:10px;padding:24px 28px;margin-bottom:20px;box-shadow:0 1px 4px rgba(0,0,0,.08)}'
'.card h2{font-size:14px;font-weight:700;color:#1F4E79;margin-bottom:16px}'
'.row{display:flex;gap:10px;align-items:flex-start;flex-wrap:wrap}'
'textarea{flex:1;min-width:280px;height:110px;border:1px solid #c8d6e5;border-radius:6px;padding:10px 12px;font-size:13px;font-family:inherit;resize:vertical;outline:none}'
'textarea:focus{border-color:#2E75B6}'
'.btn{padding:10px 22px;border:none;border-radius:6px;font-size:13px;font-weight:700;cursor:pointer;white-space:nowrap}'
'.bp{background:#1F4E79;color:#fff}.bp:hover{background:#163e61}'
'.be{background:#217346;color:#fff}.be:hover{background:#185a36}'
'.btn:disabled{opacity:.5;cursor:not-allowed}'
'.guide{font-size:12px;color:#888;margin-top:10px;line-height:1.8}'
'.status{background:#fff;border-radius:10px;padding:14px 20px;margin-bottom:16px;font-size:13px;color:#555;box-shadow:0 1px 4px rgba(0,0,0,.08);display:none}'
'.status.show{display:block}'
'.tcard{background:#fff;border-radius:10px;overflow:auto;box-shadow:0 1px 4px rgba(0,0,0,.08);margin-bottom:20px}'
'table{width:100%;border-collapse:collapse;min-width:900px}'
'th{background:#2E75B6;color:#fff;padding:10px 12px;font-size:12px;font-weight:700;text-align:center;white-space:nowrap}'
'td{padding:9px 12px;font-size:12px;border-bottom:1px solid #edf2f7;white-space:nowrap}'
'tr:nth-child(even) td{background:#f7fafd}'
'.tc{text-align:center}.tr{text-align:right}.tna{color:#bbb;text-align:center}'
'.tn{font-weight:700;color:#1F4E79}'
'.sh{text-align:center;font-size:11px;font-weight:700;padding:5px;color:#fff}'
'.empty{text-align:center;color:#aaa;padding:40px}'
'.tag{display:inline-block;font-size:10px;padding:2px 7px;border-radius:8px;font-weight:700}'
'.tok{background:#e8f5e9;color:#2e7d32}.twn{background:#fff8e1;color:#f57f17}.ter{background:#fce4ec;color:#c62828}'
'.sp{display:inline-block;width:13px;height:13px;border:2px solid #ccc;border-top-color:#1F4E79;border-radius:50%;animation:sp .7s linear infinite;vertical-align:middle;margin-right:6px}'
'@keyframes sp{to{transform:rotate(360deg)}}'
'</style></head><body>'
'<header><h1>상장사 재무정보 조회</h1>'
'<p>별도재무제표 기준 | 2025년 사업보고서 우선, 미제출시 최근 보고서로 자동 대체 | 단위: 원</p></header>'
'<div class="wrap">'
'<div class="card"><h2>기업명 입력</h2>'
'<div class="row">'
'<textarea id="ta" placeholder="기업명을 한 줄에 하나씩 입력 (최대 10개)\n예)\n고려제강\n영흥\n한국선재"></textarea>'
'<div style="display:flex;flex-direction:column;gap:8px">'
'<button class="btn bp" id="qb" onclick="doQ()">조회하기</button>'
'<button class="btn be" id="eb" onclick="dlXL()" style="display:none">엑셀 다운로드</button>'
'</div></div>'
'<div class="guide">※ 동명이인 기업이 있으면 선택 화면이 나타납니다<br>※ 종목코드로 입력해도 됩니다 (예: 005930)</div>'
'</div>'
'<div class="status" id="st"></div>'
'<div id="res"></div>'
'</div>'
'<script>'
'var LR=[];'
'function pn(r){var a=r.split("\\n"),b=[];for(var i=0;i<a.length;i++){var s=a[i].trim();if(s)b.push(s);}return b.slice(0,10);}'
'function ss(m){var e=document.getElementById("st");e.innerHTML=m;e.className="status show";}'
'function doQ(){'
'var names=pn(document.getElementById("ta").value);'
'if(!names.length){alert("기업명을 입력하세요.");return;}'
'document.getElementById("qb").disabled=true;'
'document.getElementById("eb").style.display="none";'
'document.getElementById("res").innerHTML="";'
'ss("<span class=\\"sp\\"></span>조회 중... DART 기업목록 최초 로딩시 약 10~20초 소요됩니다.");'
'fetch("/query",{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify({names:names})})'
'.then(function(r){return r.json();})'
'.then(function(d){'
'document.getElementById("qb").disabled=false;'
'if(d.error){ss("오류: "+d.error);return;}'
'if(d.candidates){showC(d.candidates,d.pending,d.done);return;}'
'LR=d.results;drawT(LR);'
'ss("완료: "+LR.length+"개 기업 조회됨");'
'document.getElementById("eb").style.display="";'
'})'
'.catch(function(e){document.getElementById("qb").disabled=false;ss("요청 실패: "+e.message);});'
'}'
'function drawT(results){'
'if(!results.length){document.getElementById("res").innerHTML="<div class=\\"empty\\">결과 없음</div>";return;}'
'var f=["매출액","매출원가","판매비와관리비","영업이익","당기순이익","자산","부채","자본"];'
'var h="<div class=\\"tcard\\"><table><thead>";'
'h+="<tr><th colspan=\\"4\\"></th><th colspan=\\"5\\" class=\\"sh\\" style=\\"background:#1a5c96\\">손익계산서</th><th colspan=\\"3\\" class=\\"sh\\" style=\\"background:#155a38\\">재무상태표</th></tr>";'
'h+="<tr><th>No.</th><th>기업명</th><th>종목코드</th><th>보고서기준</th>";'
'for(var i=0;i<f.length;i++)h+="<th>"+f[i]+"</th>";'
'h+="</tr></thead><tbody>";'
'for(var r=0;r<results.length;r++){'
'var row=results[r];'
'var tc="twn";'
'if(row["보고서기준"]==="조회실패"||row["보고서기준"]==="데이터없음")tc="ter";'
'else if(row["보고서기준"].indexOf("2025년 사업")>=0)tc="tok";'
'h+="<tr><td class=\\"tc\\">"+(r+1)+"</td><td class=\\"tn\\">"+row["기업명"]+"</td><td class=\\"tc\\">"+row["종목코드"]+"</td>";'
'h+="<td class=\\"tc\\"><span class=\\"tag "+tc+"\\">"+row["보고서기준"]+"</span></td>";'
'for(var fi=0;fi<f.length;fi++){'
'var v=row[f[fi]];'
'if(v===null||v===undefined)h+="<td class=\\"tna\\">-</td>";'
'else h+="<td class=\\"tr\\">"+Math.round(v).toLocaleString()+"</td>";'
'}'
'h+="</tr>";'
'}'
'h+="</tbody></table></div>";'
'document.getElementById("res").innerHTML=h;'
'}'
'function showC(cd,pend,done){'
'var h="<div class=\\"card\\"><h2>\'"+cd.name+"\' 검색 결과가 여러 개입니다. 선택하세요.</h2>";'
'h+="<div style=\\"display:flex;flex-wrap:wrap;gap:10px;margin-top:12px\\">";'
'for(var i=0;i<cd.list.length;i++){'
'var c=cd.list[i];'
'h+="<button class=\\"btn bp\\" style=\\"height:auto;padding:10px 16px;text-align:left\\" ";'
'h+="onclick=\\"selC(\'"+c.Code+"\',\'"+c.Name+"\',"+JSON.stringify(pend)+","+JSON.stringify(done)+")\\">";'
'h+="<div>"+c.Name+"</div><div style=\\"font-size:11px;opacity:.7\\">"+c.Code+" · "+c.Market+"</div></button>";'
'}'
'h+="<button class=\\"btn\\" style=\\"background:#eee;color:#555\\" onclick=\\"selC(null,null,"+JSON.stringify(pend)+","+JSON.stringify(done)+")\\">"+"건너뜀</button>";'
'h+="</div></div>";'
'document.getElementById("res").innerHTML=h;'
'}'
'function selC(code,name,pend,done){'
'var sel=code?[{code:code,name:name}]:[];'
'ss("<span class=\\"sp\\"></span>계속 조회 중...");'
'document.getElementById("res").innerHTML="";'
'fetch("/query_continue",{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify({selected:sel,pending:pend,done:done})})'
'.then(function(r){return r.json();})'
'.then(function(d){'
'if(d.candidates){showC(d.candidates,d.pending,d.done);return;}'
'LR=d.results;drawT(LR);'
'ss("완료: "+LR.length+"개 기업 조회됨");'
'document.getElementById("eb").style.display="";'
'})'
'.catch(function(e){ss("오류: "+e.message);});'
'}'
'function dlXL(){'
'if(!LR.length)return;'
'fetch("/download",{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify({results:LR})})'
'.then(function(r){return r.blob();})'
'.then(function(b){'
'var u=URL.createObjectURL(b);var a=document.createElement("a");'
'a.href=u;a.download="재무정보_"+new Date().toISOString().slice(0,10)+".xlsx";a.click();URL.revokeObjectURL(u);'
'});'
'}'
'</script></body></html>'
)

@app.route("/")
def index():
    return HTML

@app.route("/query", methods=["POST"])
def query():
    try:
        names = request.json.get("names", [])
        sl = get_stock_list()
        cm = build_corp_map()
        resolved = []
        for name in names:
            code, rname, cands = resolve_code(name, sl)
            if cands:
                remaining = [n for n in names if n not in [name]+[r["name"] for r in resolved]]
                return jsonify({"candidates":{"name":name,"list":cands},"pending":remaining,"done":resolved})
            if code:
                resolved.append({"code":code,"name":rname})
        results = []
        for r in resolved:
            results.append(query_one(r["code"],r["name"],cm))
            time.sleep(0.3)
        return jsonify({"results":results})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/query_continue", methods=["POST"])
def query_continue():
    try:
        body = request.json
        selected = body.get("selected",[])
        pending  = body.get("pending",[])
        done     = body.get("done",[])
        sl = get_stock_list()
        cm = build_corp_map()
        resolved = list(done) + selected
        for name in list(pending):
            code, rname, cands = resolve_code(name, sl)
            if cands:
                remaining = [n for n in pending if n != name]
                return jsonify({"candidates":{"name":name,"list":cands},"pending":remaining,"done":resolved})
            if code:
                resolved.append({"code":code,"name":rname})
        results = []
        for r in resolved:
            results.append(query_one(r["code"],r["name"],cm))
            time.sleep(0.3)
        return jsonify({"results":results})
    except Exception as e:
        return jsonify({"error":str(e)}), 500

@app.route("/download", methods=["POST"])
def download():
    results = request.json.get("results",[])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "재무정보"
    tf=Font(name="Arial",bold=True,size=13,color="FFFFFF")
    tfi=PatternFill("solid",start_color="1F4E79")
    hf=Font(name="Arial",bold=True,size=10,color="FFFFFF")
    hfi=PatternFill("solid",start_color="2E75B6")
    sf=Font(name="Arial",size=9,color="808080")
    bf=Font(name="Arial",size=10)
    nf=Font(name="Arial",size=10,color="AAAAAA")
    ca=Alignment(horizontal="center",vertical="center",wrap_text=True)
    la=Alignment(horizontal="left",vertical="center")
    ra=Alignment(horizontal="right",vertical="center")
    bn=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
    af=PatternFill("solid",start_color="F2F7FC")
    cols=["No.","기업명","종목코드","보고서기준"]+list(TARGET_ACCOUNTS.keys())
    cws=[5,16,11,16,17,17,17,14,14,17,14,13]
    for i,w in enumerate(cws,start=1):
        ws.column_dimensions[chr(64+i)].width=w
    ws.row_dimensions[1].height=8
    ws.row_dimensions[2].height=32
    ws.row_dimensions[3].height=16
    ws.row_dimensions[4].height=30
    lc=chr(64+len(cols))
    ws.merge_cells(f"A2:{lc}2")
    ws["A2"]="상장사 재무정보 일괄 조회  |  별도재무제표 기준"
    ws["A2"].font=tf; ws["A2"].fill=tfi; ws["A2"].alignment=ca
    ws.merge_cells(f"A3:{lc}3")
    ws["A3"]=f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  단위: 원"
    ws["A3"].font=sf; ws["A3"].alignment=ca
    for ci,cn in enumerate(cols,start=1):
        c=ws.cell(row=4,column=ci,value=cn)
        c.font=hf; c.fill=hfi; c.alignment=ca; c.border=bn
    for ri,rd in enumerate(results,start=5):
        ws.row_dimensions[ri].height=20
        fill=af if ri%2==0 else None
        vals=[ri-4,rd.get("기업명"),rd.get("종목코드"),rd.get("보고서기준")]+[rd.get(k) for k in TARGET_ACCOUNTS]
        for ci,val in enumerate(vals,start=1):
            c=ws.cell(row=ri,column=ci)
            c.border=bn
            if fill: c.fill=fill
            if ci<=4:
                c.value=val if val is not None else "N/A"
                c.font=bf; c.alignment=ca if ci in(1,3) else la
            else:
                if val is not None:
                    c.value=val; c.number_format="#,##0"; c.font=bf; c.alignment=ra
                else:
                    c.value="N/A"; c.font=nf; c.alignment=ca
    ws.freeze_panes="E5"
    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,download_name="재무정보.xlsx",as_attachment=True,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__=="__main__":
    port=int(os.environ.get("PORT",5000))
    app.run(host="0.0.0.0",port=port)
